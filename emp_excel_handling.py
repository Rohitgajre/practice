import openpyxl

from openpyxl import Workbook, load_workbook

FILE_PATH = r"D:\Python_Course\Excel_file_Handling\Excel_Project\emp_data.xlsx"

def create_wb_sheet():
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Emp_Data'
    wb.save(FILE_PATH)

def get_wb_sheet():
    wb = load_workbook(FILE_PATH)
    sheet = wb["Emp_Data"]
    return wb, sheet


if __name__ == "__main__":

    create_wb_sheet()
    # get_wb_sheet()