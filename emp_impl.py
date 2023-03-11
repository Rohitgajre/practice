
from openpyxl.styles import Alignment
from emp_abstraction import EmpAbstraction
from emp_class import Address, Employee
from emp_data import generate_emps
from emp_excel_handling import FILE_PATH, create_wb_sheet, get_wb_sheet
import openpyxl
from openpyxl import load_workbook



class EmpImplementation(EmpAbstraction):
    def align_cells(self, sheet):
        for row in range(1, sheet.max_row+1):
            for col in range(1, sheet.max_column+1):
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    def write_headers(self):
        wb, sheet = get_wb_sheet()
        headers = "ID Name Age Salary Address".split()
        for i in range(1, 6):    
            sheet.cell(row=1, column=i).value = headers[i-1]  
            sheet.merge_cells('E1:G1')
        adr_headers = "Pincode City State".split()
        for i in range(5, 8):  
            sheet.cell(row=2, column=i).value = adr_headers[i-5]  
        wb.save(FILE_PATH)

    def write_data(self):
        wb, sheet = get_wb_sheet()
        all_emps = generate_emps(100)
        row = 3           
        for emp in all_emps: 
            # if emp.EmpSalary <= 25000:
            #     emp.EmpSalary + emp.EmpSalary * 5/100   
                sheet.cell(row=row, column=1).value = emp.EmpID
                sheet.cell(row=row, column=2).value = emp.EmpName
                sheet.cell(row=row, column=3).value = emp.EmpAge
                sheet.cell(row=row, column=4).value = emp.EmpSalary
                sheet.cell(row=row, column=5).value = emp.EmpAddress[0].Pincode
                sheet.cell(row=row, column=6).value = emp.EmpAddress[0].City
                sheet.cell(row=row, column=7).value = emp.EmpAddress[0].State
                row += 1        
        self.align_cells(sheet)
        try:
            wb.save(FILE_PATH)
        except PermissionError:
            print(f"File is open. Please close the file before write opearation:- {FILE_PATH}")
     


    def read_data(self):
        
        wb, sheet = get_wb_sheet()
        list_of_emps = []
        for i in range(3, sheet.max_row+1):           
                eid = sheet.cell(row=i, column=1).value
                ename = sheet.cell(row=i, column=2).value
                eage = sheet.cell(row=i, column=3).value
                esalary = sheet.cell(row=i, column=4).value
                pin = sheet.cell(row=i, column=5).value
                city = sheet.cell(row=i, column=6).value
                state = sheet.cell(row=i, column=7).value
                emp_obj = Employee(eid=eid, ename=ename, eage=eage, esal=esalary)
                adr_obj = Address(pin=pin, city=city, state=state)
                emp_obj.EmpAddress.append(adr_obj)
                list_of_emps.append(emp_obj)
        return list_of_emps

        
        
# ---------------particilar emp by id------------------------------------------------------------
  
    def employee_id(self):
        list_emp = self.read_data()
        result = []
        Number = int(input("Enter employee id:-"))
        for emp in list_emp:
            if emp.EmpID == Number:
                result.append(emp)
                # print(emp)
                return emp

#-------------------------# salary is less than -- 25000 ==> 5% increment and write data in excel---------------------------    

    def emp_by_sal(self):
        wb, sheet = get_wb_sheet()
        increment_sal = []
        for i in range(3, sheet.max_row+1):           
            esalary = sheet.cell(row=i, column=4).value
            if esalary <= 25000:
                EmpSalary = esalary + esalary * 5/100 
                increment_sal.append(EmpSalary)
        # print("Employees list :- ", update_sal)
        wb.save(FILE_PATH)
        return increment_sal
               
# ----------------# first n number of records------------- done
    def first_n_number(self):
        list_of_emps = self.read_data()
        Number_list = []
        emps = int(input("Enter the employee count :-"))
        for i in range(0 , emps):
            Number_list.append(list_of_emps[i])
        return Number_list
# -------------# record in between 10-125 id kiva -----------------
    def get_between(self):
        Number_list = []
        list_of_emps = self.read_data()
        print(type(list_of_emps))
        emps = int(input("Enter the number of sheet record from:-"))
        emp = int(input("Enter to from :-"))
        for i in range(emps , emp):
            Number_list.append(list_of_emps[i])
        return Number_list

        
  
if __name__ == "__main__":
    obj = EmpImplementation()
    # obj.write_headers()
    # obj.write_data() 
    # print(obj.read_data())

# assignment operations :-

    # print(obj.employee_id())                              #  particilar emp by id-
    print(obj.emp_by_sal())                               #  salary is less than -- 25000 ==> 5% increment and write data in excel
    # print(obj.first_n_number())                           #  first n number of records
    # print(obj.get_between())                              #  record in between 105-125 id kiva

    


